#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""Creates insightful structured ata based on provided spreadsheet's contents."""

__author__ = 'Diego Rocha'
__maintainer__ = __author__
__email__ = 'dhsrocha.dev@gmail.com'
__version__ = '0.1.0-SNAPSHOT'

import abc
import csv
import os
import re
import sys
from abc import abstractmethod
from datetime import datetime as dt
from logging import Logger, getLogger
from pathlib import Path
from random import randint
from tempfile import TemporaryDirectory
from typing import List, Any, Callable, Type
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class _Analyzer:
    """
    Nodule's namespace class.

    Meant to enclose module's resources and protect them from outside scope.
    """

    # TODO Implement GUI, under switch case structure

    _csv_ext: str = '.csv'
    _pwd: Path = Path(os.path.dirname(os.path.realpath(__file__)))
    _out_json: Path = _pwd.joinpath('output', 'json')
    _out_csv: Path = _pwd.joinpath('output', 'csv')
    _re_xls: re.Pattern[str] = re.compile('.+\\.xls[xm]')

    @staticmethod
    def _throw[T](ex: Type[T], message: str) -> None:
        """
        Utility method to check and throw exception in one line.

        :param ex: the exception type to be thrown.
        :param message: the exception message.
        :raises T: the provided exception type.
        """
        raise ex(message)

    class _Loggable(metaclass=abc.ABCMeta):
        """Abstraction for objects that can be logged. """

        _log: Logger

        @abstractmethod
        def __init__(self):
            """
            Constructor function to set up a named logger instance based on type's name.
            """
            self._log: Logger = getLogger(self.__class__.__name__.replace('_', ''))

    # ::: Back-End

    class _BackEnd(metaclass=abc.ABCMeta):
        """Abstraction for the application's back-end layer."""

        @abstractmethod
        def normalize(self,
                      dir_source: Path,
                      dir_destination: Path,
                      overwrites_destination: bool) -> None:
            """
            Structure content from spreadsheet files to csv.

            A self-test example with temporary files and folder:

            :param dir_source: Directory containing the spreadsheets to ingest contents
                               from.
            :type dir_source: Path
            :param dir_destination: Destination directory for the parsed content,
                                    sub-folders wilt be created if it does not exist.
            :type dir_destination: Path
            :param overwrites_destination: Overwrites destination_dir's contents if True
            :type overwrites_destination: bool
            :raises ValueError: if source_dir does not exist.
            """
            pass

    class _ToCsv(_Loggable, _BackEnd):
        """Produces normalized content from inputs."""

        def __init__(self):
            """Required to call super()."""
            super().__init__()

        def normalize(self,
                      dir_source: Path,
                      dir_destination: Path,
                      overwrites_destination: bool) -> None:
            """
            >>> suite = _Analyzer._ToCsv._Tests
            >>> suite.given_random_content__when_run_then_nothing_is_raised()
            """
            _start = dt.now()
            # validate input
            if not dir_source.is_dir() or not dir_source.exists():
                _Analyzer._throw(_Analyzer._InputException,
                                 f'Parameter for input folder '
                                 f'"{dir_source}" is not a directory."]')
                self._log.debug(f'Input directory to use: [{dir_source}].')

            # Create or validate output
            if not dir_destination.exists():
                _out: Path = Path(os.path.join(os.getcwd(), dir_destination))
                os.makedirs(_out)
                if overwrites_destination:
                    [os.remove(_out.joinpath(f)) for f in os.listdir(_out)]
                    os.rmdir(_out)
                    self._log.warning(f'{_out}\'s contents removed.')
                self._log.warning(f'Output folder "{_out}" created.')
            self._log.debug(f'Output directory to use: [{dir_destination}].')

            # Load worksheets from xls in the input folder
            _worksheets: dict[Worksheet, str] = dict()
            for _p in dir_source.iterdir():
                if _p.is_file() and _Analyzer._re_xls.match(str(_p)):
                    self._log.debug(f'Opened spreadsheet: [{_p.stem}{_p.suffix}].')
                    _name: str = re.sub('[$~]', '', str(_p))
                    _worksheets.update([(ws, Path(_name).stem) for ws in
                                        load_workbook(filename=_name, data_only=True)])
                    self._log.info(
                        f'Loaded [{len(_worksheets)}] worksheets from files '
                        f'in "{dir_source}".')

            # TODO Strip trailing and leading spaces and linebreaks
            # TODO Convert non-utf8 characters

            _cleanse: Callable[[Any], str] = lambda v: v.replace('\n', '\\n').strip() \
                if type(v) is str else v

            # Write cleaned content into csv files
            for _ws, _filename in _worksheets.items():
                _concat: str = (_ws.title if _ws.title is _filename
                                else f'{_filename}_{_ws.title}') + _Analyzer._csv_ext
                _cells: list[tuple] = list(zip(*[col for col in _ws.columns
                                                 if any(c.value is not None for c in
                                                        col)]))
                self._log.debug(f'"{_concat}" contents parsed.')
                with open(dir_destination.joinpath(_concat), 'w+',
                          newline='') as _csv_io:
                    _contents: List[list] = [[_cleanse(c.value) for c in r]
                                             for r in _cells]
                    csv.writer(_csv_io).writerows(_contents)
                self._log.debug(f'"[_ws]" written to "{_concat}".')
            self._log.info('Spreadsheet structured to CSV in '
                           f'"{dir_destination}" in [{dt.now() - _start}].')

        class _Tests:
            """Suite for system tests."""

            # TODO Turn tests repeatable and maybe parametric

            @staticmethod
            def given_random_content__when_run_then_nothing_is_raised() -> None:
                f"""GIVEN temporary xls file\n
                WHEN convert created file from it\n
                THEN written {_Analyzer._csv_ext} in output as per worksheet\n
                AND each one has content as provided.
                """
                with (TemporaryDirectory() as _root):
                    # Given
                    _out: Path = Path(_root).joinpath(_Analyzer._out_csv)
                    _name: str = str(uuid4())
                    _ws1: str = str(uuid4())[:31]
                    _ws2: str = str(uuid4())[:31]

                    _wb: Workbook = Workbook()
                    _wb.create_sheet(_ws1)
                    _wb.create_sheet(_ws2)
                    _wb.remove(_wb['Sheet'])

                    [_wb[_ws2].append(r) for r in
                     _Analyzer._ToCsv._Tests._full_contents()]
                    [_wb[_ws1].append(r) for r in
                     _Analyzer._ToCsv._Tests._full_contents()]
                    _wb.save(Path(_root).joinpath(_name + '.xlsx'))

                    # When
                    _Analyzer._ToCsv().normalize(Path(_root), _out, False)

                    # Then
                    _out_files = os.listdir(_out)
                    assert len(_out_files) == len(_wb.sheetnames), \
                        f'"{_out}" must not be empty.'

                    for _ws in _wb.worksheets:
                        _nm = f'{_name}_{_ws.title}_{_Analyzer._csv_ext}'
                        with open(_out.joinpath(_nm)) as _csv_io:
                            _reader: csv.DictReader = csv.DictReader(_csv_io)
                            _extract: list = [[_reader.fieldnames] + [row.values()] for
                                              row in _reader]
                        _expect: list = [[str(cel.value).replace('\n', '\\n')
                                          for cel in row] for row in _ws.rows]

                        assert len(_expect) == len(_extract), \
                            f'Row size: [{len(_expect)}] not same as [{len(_extract)}],'

                        assert _expect == _extract, \
                            f'Contents: {_expect} not sane as {_extract}.'

                    # TODO Create test case for not coinciding worksheet names and
                    #  its corresponding filenames
                    # TODO Create test case for rows with cells with blank/empty/None
                    #  values
                    # TODO Create test case for irregular / non-rectangular dato-frames

            @classmethod
            def _full_contents(cls) -> List[dict[str, str]]:
                """Helper method to dynamically generate contents."""

                def cell() -> str: return f'\n{uuid4()}\n'

                def row() -> dict[str, str]: return dict(A=cell(), B=cell(), C=cell(), )

                return [row() for _ in range(randint(20, 100))]

    # Exceptions

    class _InputException(Exception):
        pass

    if __name__ == '__main__':
        sys.version_info < (3, 12) and _throw(RuntimeError, 'Minimum version is 3.12')
        print('Hello World!')
